import streamlit as st
import pdfplumber
import pandas as pd
import re
import io
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# v2.5 - fix: label jednoznakowy z myslnikiem (-V, -W), podwojny myslnik w stylu
st.set_page_config(page_title="PO Extractor", page_icon="📦", layout="centered")

st.title("📦 PO Extractor")

# ── PARSERY ────────────────────────────────────────────────────────────────────

LEVIS_PATTERN = re.compile(
    r'^\d{10}'
    r'([A-Z0-9]{1,10}-[A-Z0-9]{2,4}(?:-[A-Z0-9]{0,6})?)'
    r'.+?'
    r'\s+(\d+)\s+EA\s+'
    r'(\d+\.\d{2})'
)

NIKE_LINE_PATTERN = re.compile(
    r'^\d{1,2}\s+'
    r'([A-Z0-9]{6}-[A-Z0-9]{3}(?:-{1,2}[A-Z0-9]{1,2})?)'  # -{1,2} obsługuje -- przed labelem
    r'\s+.+?'
    r'\s+\$(\d+\.\d{2})'
    r'\s+(\d+)'
    r'\s+\$[\d,]+\.\d{2}'
)

BABYLIST_PATTERN = re.compile(
    r'^(BL-\d+)'
    r'\s*\|\s*'
    r'([A-Z0-9]{5,8})\s*-\s*([A-Z0-9]{3})'
    r'(?:\s*-\s*([A-Z0-9]{2}))?'
    r'\s*\|.+?'
    r'\s+\$(\d+\.\d{2})'
    r'\s+(\d+)'
    r'\s+\$[\d,]+\.\d{2}$'
)

def extract_levis(pdf_file):
    rows = []
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text: continue
            for line in text.split("\n"):
                line = line.strip()
                m = LEVIS_PATTERN.match(line)
                if m:
                    rows.append({
                        "Vendor Style ID": m.group(1),
                        "Total Units": int(m.group(2)),
                        "Vendor First Cost (USD)": float(m.group(3)),
                        "SKU": "",
                    })
    return rows

def extract_nike(pdf_file):
    rows = []
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text: continue
            for line in text.split("\n"):
                line = line.strip()
                m = NIKE_LINE_PATTERN.match(line)
                if m:
                    rows.append({
                        "Vendor Style ID": m.group(1),
                        "Total Units": int(m.group(3)),
                        "Vendor First Cost (USD)": float(m.group(2)),
                        "SKU": "",
                    })
    return rows

def extract_sierra(pdf_file):
    """Parser dla S894M (Sierra/TJX). Sumuje unity per styl, rozdziela 0810/0860.
    Dwa przebiegi: najpierw ceny (przed tabelką DC), potem unity (tabelki DC).
    Działa poprawnie gdy ceny i tabelka DC są na tej samej stronie."""
    from collections import defaultdict
    style_prices = {}
    style_units  = defaultdict(lambda: {'units_0810': 0, 'units_0860': 0})

    STYLE_PAGE1 = re.compile(r'^\d{1,2}/\d{1,2}/\d{4}\s+([A-Z0-9]{6}-[A-Z0-9]{3}(?:-{1,2}[A-Z0-9]{1,2})?),?')
    PRICE_LINE  = re.compile(r'^\d[\d,]*\s+\$(\d+\.\d{2})\s+\$[\d,]+\.\d{2}$')
    DC_ROW      = re.compile(r'^([A-Z0-9]{6}-[A-Z0-9]{3}(?:-{1,2}[A-Z0-9]{1,2})?),?\s+\S+\s+.+?\s+\S+\s+(\d+)\s+(\d+)\s+(\d+)\s*$')
    DC_HEADER   = re.compile(r'^Vendor Styles\s+Item Code')

    all_lines = []
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                all_lines.extend([l.strip() for l in text.split("\n") if l.strip()])

    # Znajdź gdzie zaczyna się pierwsza tabelka DC
    dc_start_idx = next((i for i, l in enumerate(all_lines) if DC_HEADER.match(l)), None)

    # Przebieg 1: ceny TYLKO przed tabelką DC
    current_style = None
    for line in (all_lines[:dc_start_idx] if dc_start_idx else all_lines):
        m = STYLE_PAGE1.match(line)
        if m: current_style = m.group(1)
        m = PRICE_LINE.match(line)
        if m and current_style:
            style_prices[current_style] = float(m.group(1))
            current_style = None

    # Przebieg 2: unity z tabelek DC (kontynuuje przez kolejne strony)
    in_dc = False
    for line in all_lines:
        if DC_HEADER.match(line):
            in_dc = True
            continue
        if in_dc:
            m = DC_ROW.match(line)
            if m:
                vid = m.group(1)
                style_units[vid]['units_0810'] += int(m.group(3))
                style_units[vid]['units_0860'] += int(m.group(4))

    rows = []
    for vid, units in style_units.items():
        rows.append({
            'Vendor Style ID': vid,
            'Total Units': units['units_0810'] + units['units_0860'],
            'Vendor First Cost (USD)': style_prices.get(vid, 0.0),
            'SKU': '',
            'units_0810': units['units_0810'],
            'units_0860': units['units_0860'],
        })
    return rows

def extract_babylist(pdf_file):
    rows = []
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text: continue
            for line in text.split("\n"):
                line = line.strip()
                m = BABYLIST_PATTERN.match(line)
                if m:
                    rows.append({
                        "Vendor Style ID": f"{m.group(2)}-{m.group(3)}" + (f"-{m.group(4)}" if m.group(4) else ""),
                        "Total Units": int(m.group(6)),
                        "Vendor First Cost (USD)": float(m.group(5)),
                        "SKU": m.group(1),
                    })
    return rows

ACCOUNT_PARSER = {
    "C810M": extract_levis,
    "B614M": extract_babylist,
    "M004M": extract_nike,
    "S894M": extract_sierra,
}

# ── HELPERS ────────────────────────────────────────────────────────────────────

def split_style_id(vendor_style_id):
    # Obsługuje zarówno pojedynczy myślnik (86F633-023-WH)
    # jak i podwójny myślnik (86M344-U90--V)
    import re as _re
    m = _re.match(r'^([A-Z0-9]+)-([A-Z0-9]+)-{1,2}([A-Z0-9]{1,2})$', vendor_style_id)
    if m:
        label = ('-' + m.group(3)) if len(m.group(3)) == 1 else m.group(3)
        return m.group(1), m.group(2), label
    parts = vendor_style_id.split('-')
    style = parts[0] if len(parts) > 0 else ''
    color = parts[1] if len(parts) > 1 else ''
    label = ''
    return style, color, label

def write_excel_sierra(rows):
    """Excel dla S894M - szablon MultistoreCart z kolumnami G=0810, H=0860."""
    TEMPLATE_PATH = 'In-LineCart_template.xlsx'
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb['Sheet1']
    # Kolumny: A=Div, B=Style#, C=Color, D=Label, E=Price, F=SKU, G=0810 Units, H=0860 Units
    start_row = 2
    for i, row in enumerate(rows):
        style, color, label = split_style_id(row['Vendor Style ID'])
        r = start_row + i
        ws.cell(row=r, column=1).value = ''
        ws.cell(row=r, column=2).value = style
        ws.cell(row=r, column=3).value = color
        ws.cell(row=r, column=4).value = label
        ws.cell(row=r, column=5).value = row['Vendor First Cost (USD)']
        ws.cell(row=r, column=6).value = row.get('SKU', '')
        ws.cell(row=r, column=7).value = row.get('units_0810', 0)
        ws.cell(row=r, column=8).value = row.get('units_0860', 0)

    total_row = start_row + len(rows)
    total_0810 = sum(r.get('units_0810', 0) for r in rows)
    total_0860 = sum(r.get('units_0860', 0) for r in rows)
    total_cost = sum(r['Total Units'] * r['Vendor First Cost (USD)'] for r in rows)
    ws.cell(row=total_row, column=2).value = 'TOTAL'
    ws.cell(row=total_row, column=5).value = round(total_cost, 2)
    ws.cell(row=total_row, column=7).value = total_0810
    ws.cell(row=total_row, column=8).value = total_0860

    bold = Font(bold=True)
    fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col_idx in [2, 5, 7, 8]:
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def write_excel_sierra_plain(rows):
    """Fallback bez szablonu dla S894M."""
    out_data = []
    for row in rows:
        style, color, label = split_style_id(row['Vendor Style ID'])
        out_data.append({
            'Div': '',
            'Style#': style,
            'Color': color,
            'Label': label,
            'Price': row['Vendor First Cost (USD)'],
            'SKU': '',
            '': row.get('units_0810', 0),
            '3': row.get('units_0860', 0),
        })
    df = pd.DataFrame(out_data)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        ws = writer.sheets["Sheet1"]
        total_row = len(df) + 2
        ws.cell(row=total_row, column=2).value = 'TOTAL'
        ws.cell(row=total_row, column=5).value = round(sum(r['Total Units']*r['Vendor First Cost (USD)'] for r in rows), 2)
        ws.cell(row=total_row, column=7).value = sum(r.get('units_0810', 0) for r in rows)
        ws.cell(row=total_row, column=8).value = sum(r.get('units_0860', 0) for r in rows)
        bold = Font(bold=True)
        fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for col_idx in [2, 5, 7, 8]:
            cell = ws.cell(row=total_row, column=col_idx)
            cell.font = bold
            cell.fill = fill
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4
    buffer.seek(0)
    return buffer

def write_excel_template(rows):
    TEMPLATE_PATH = 'In-LineCart_template.xlsx'
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb['Sheet1']
    start_row = 2
    for i, row in enumerate(rows):
        style, color, label = split_style_id(row['Vendor Style ID'])
        r = start_row + i
        ws.cell(row=r, column=1).value = ''
        ws.cell(row=r, column=2).value = style
        ws.cell(row=r, column=3).value = color
        ws.cell(row=r, column=4).value = label
        ws.cell(row=r, column=5).value = row['Vendor First Cost (USD)']
        ws.cell(row=r, column=6).value = row.get('SKU', '')
        ws.cell(row=r, column=7).value = row['Total Units']

    total_row = start_row + len(rows)
    total_units = sum(r['Total Units'] for r in rows)
    total_cost = sum(r['Total Units'] * r['Vendor First Cost (USD)'] for r in rows)
    ws.cell(row=total_row, column=2).value = 'TOTAL'
    ws.cell(row=total_row, column=5).value = round(total_cost, 2)
    ws.cell(row=total_row, column=7).value = total_units
    bold = Font(bold=True)
    fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col_idx in [2, 5, 7]:
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def write_excel_plain(rows):
    df = pd.DataFrame(rows)
    df['style'] = df['Vendor Style ID'].apply(lambda x: split_style_id(x)[0])
    df['color'] = df['Vendor Style ID'].apply(lambda x: split_style_id(x)[1])
    df['label'] = df['Vendor Style ID'].apply(lambda x: split_style_id(x)[2])
    out = pd.DataFrame({
        'Div': '',
        'Style#': df['style'],
        'Color': df['color'],
        'Label': df['label'],
        'Cost': df['Vendor First Cost (USD)'],
        'SKU': df['SKU'] if 'SKU' in df.columns else '',
        'Pcs': df['Total Units'],
    })
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="Sheet1")
        ws = writer.sheets["Sheet1"]
        total_row = len(out) + 2
        ws.cell(row=total_row, column=2).value = 'TOTAL'
        ws.cell(row=total_row, column=5).value = round(df['Total Units'].mul(df['Vendor First Cost (USD)']).sum(), 2)
        ws.cell(row=total_row, column=7).value = int(df['Total Units'].sum())
        bold = Font(bold=True)
        fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for col_idx in [2, 5, 7]:
            cell = ws.cell(row=total_row, column=col_idx)
            cell.font = bold
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4
    buffer.seek(0)
    return buffer

# ── UI ─────────────────────────────────────────────────────────────────────────

# Inicjalizacja stanu
if "selected_account" not in st.session_state:
    st.session_state.selected_account = None

# Kafelki kont
st.markdown("### Wybierz konto")

ACCOUNTS = [
    {"name": "C810M", "color": "#1f4e79", "desc": "Century 21 / Haddad"},
    {"name": "B614M", "color": "#375623", "desc": "Babylist / Huggies"},
    {"name": "M004M", "color": "#7b2c2c", "desc": "Macy's Backstage"},
    {"name": "S894M", "color": "#4a235a", "desc": "Sierra / TJX"},
]

cols = st.columns(len(ACCOUNTS))
for col, acc in zip(cols, ACCOUNTS):
    with col:
        is_selected = st.session_state.selected_account == acc["name"]
        border = "4px solid #FFD700" if is_selected else "2px solid transparent"
        st.markdown(
            f"""
            <div style="
                background-color: {acc['color']};
                border: {border};
                border-radius: 12px;
                padding: 20px 10px;
                text-align: center;
                margin-bottom: 8px;
            ">
                <div style="color: white; font-size: 22px; font-weight: bold;">{acc['name']}</div>
                <div style="color: #ccc; font-size: 12px; margin-top: 4px;">{acc['desc']}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button(
            "✓ Wybrano" if is_selected else "Wybierz",
            key=f"btn_{acc['name']}",
            use_container_width=True,
        ):
            st.session_state.selected_account = acc["name"]
            st.rerun()

# Pokaż resztę UI tylko gdy konto jest wybrane
if st.session_state.selected_account:
    account = st.session_state.selected_account
    st.success(f"✅ Wybrane konto: **{account}**")
    st.divider()

    uploaded_file = st.file_uploader("Wybierz plik PDF", type="pdf")

    template_file = st.file_uploader(
        "📋 Wgraj szablon Excel (In-LineCart) — opcjonalnie",
        type=["xlsx"],
        help="Jeśli wgrasz szablon, dane zostaną wstawione do Sheet1."
    )

    if uploaded_file is not None:
        with st.spinner("Przetwarzam PDF..."):
            try:
                parser = ACCOUNT_PARSER[account]
                rows = parser(uploaded_file)

                if not rows:
                    st.error("Nie udało się wyciągnąć danych. Sprawdź czy wgrałeś właściwy plik dla tego konta.")
                else:
                    preview = []
                    for row in rows:
                        style, color, label = split_style_id(row['Vendor Style ID'])
                        preview.append({
                            'Style#': style,
                            'Color': color,
                            'Label': label,
                            'Cost': row['Vendor First Cost (USD)'],
                            'SKU': row.get('SKU', ''),
                            'Pcs': row['Total Units'],
                        })
                    df_preview = pd.DataFrame(preview)

                    st.success(f"✅ Wyciągnięto **{len(rows)}** pozycji z PDF.")
                    st.dataframe(df_preview, use_container_width=True)

                    col1, col2 = st.columns(2)
                    total_units = sum(r['Total Units'] for r in rows)
                    total_cost = sum(r['Total Units'] * r['Vendor First Cost (USD)'] for r in rows)
                    col1.metric("Łączna liczba jednostek", f"{total_units:,}")
                    col2.metric("Łączny koszt (USD)", f"${total_cost:,.2f}")

                    if account == "S894M":
                        if template_file is not None:
                            template_bytes = template_file.read()
                            with open('In-LineCart_template.xlsx', 'wb') as f:
                                f.write(template_bytes)
                            buffer = write_excel_sierra(rows)
                            filename = "MultistoreCart_filled.xlsx"
                        else:
                            buffer = write_excel_sierra_plain(rows)
                            filename = "po_data_sierra.xlsx"
                    elif template_file is not None:
                        template_bytes = template_file.read()
                        with open('In-LineCart_template.xlsx', 'wb') as f:
                            f.write(template_bytes)
                        buffer = write_excel_template(rows)
                        filename = "In-LineCart_filled.xlsx"
                    else:
                        buffer = write_excel_plain(rows)
                        filename = "po_data.xlsx"

                    st.download_button(
                        label="⬇️ Pobierz Excel",
                        data=buffer,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

            except Exception as e:
                st.error(f"Błąd podczas przetwarzania: {e}")
                st.exception(e)
else:
    st.info("👆 Najpierw wybierz konto powyżej.")
